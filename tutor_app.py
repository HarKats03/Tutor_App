import customtkinter as ctk
import sqlite3
from datetime import datetime
import pandas as pd
import os
import platform
import calendar
import shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageDraw, ImageTk

# --- ΕΙΔΙΚΗ ΡΥΘΜΙΣΗ ΓΙΑ ΤΗΝ TASKBAR ΤΩΝ WINDOWS ---
if platform.system() == 'Windows':
    import ctypes

    try:
        app_id = 'tutor.manager.pro.final.v8'
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(app_id)
    except Exception:
        pass


# --- 1. ΑΥΤΟΜΑΤΗ ΔΗΜΙΟΥΡΓΙΑ ΛΟΓΟΤΥΠΟΥ ---
def ensure_logo_exists():
    if not os.path.exists("logo.png") or not os.path.exists("logo.ico"):
        size = (512, 512)
        img = Image.new('RGBA', size, (255, 255, 255, 0))
        draw = ImageDraw.Draw(img)
        x0, y0, x1, y1 = 32, 32, 480, 480
        draw.rounded_rectangle((x0, y0, x1, y1), radius=110, fill="#0A84FF")
        draw.rectangle([216, 140, 296, 400], fill="white")
        draw.rectangle([130, 140, 382, 210], fill="white")
        img.save("logo.png")
        img.save("logo.ico", format='ICO', sizes=[(256, 256), (128, 128), (64, 64), (32, 32)])


def apply_window_icon(window):
    try:
        if platform.system() == "Windows" and os.path.exists("logo.ico"):
            window.iconbitmap("logo.ico")
        else:
            icon_img = ImageTk.PhotoImage(Image.open("logo.png"))
            window.iconphoto(True, icon_img)
    except Exception:
        pass


# --- 2. CUSTOM ALERT DIALOG ---
def show_custom_alert(parent, title, message, is_error=False, callback=None):
    dialog = ctk.CTkToplevel(parent)
    dialog.title(title)
    dialog.geometry("450x250")
    dialog.transient(parent)
    dialog.grab_set()

    dialog.update_idletasks()
    x = parent.winfo_x() + (parent.winfo_width() // 2) - (450 // 2)
    y = parent.winfo_y() + (parent.winfo_height() // 2) - (250 // 2)
    dialog.geometry(f"+{x}+{y}")

    logo_img = ctk.CTkImage(Image.open("logo.png"), size=(60, 60))
    ctk.CTkLabel(dialog, text="", image=logo_img).pack(pady=(20, 10))

    txt_color = "#e74c3c" if is_error else "#ffffff"
    ctk.CTkLabel(dialog, text=title, font=("Arial", 18, "bold"), text_color=txt_color).pack(pady=5)
    ctk.CTkLabel(dialog, text=message, font=("Arial", 14), wraplength=400).pack(pady=5)

    def on_click():
        dialog.destroy()
        if callback: callback()

    btn_color = "#e74c3c" if is_error else "#27ae60"
    ctk.CTkButton(dialog, text="ΟΚ", fg_color=btn_color, width=120, command=on_click).pack(pady=15)


# --- 3. ΒΑΣΗ ΔΕΔΟΜΕΝΩΝ ---
def init_db():
    with sqlite3.connect("tutor_manager.db") as conn:
        cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS students (
                            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, 
                            group_name TEXT, rate_per_hour REAL NOT NULL, hours_per_session REAL NOT NULL)''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS schedule (
                            id INTEGER PRIMARY KEY AUTOINCREMENT, student_id INTEGER, 
                            day_of_week TEXT NOT NULL, FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE)''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS session_logs (
                            id INTEGER PRIMARY KEY AUTOINCREMENT, student_id INTEGER, 
                            date TEXT NOT NULL, hours_done REAL NOT NULL, earned_amount REAL NOT NULL, 
                            FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE)''')
        try:
            cursor.execute("ALTER TABLE session_logs ADD COLUMN notes TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        conn.commit()


# --- 4. Η ΚΥΡΙΑ ΕΦΑΡΜΟΓΗ ---
class TutorApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.attributes("-alpha", 0.0)
        self.withdraw()

        self.title("Tutor Manager Pro+")
        self.geometry("1400x850")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        apply_window_icon(self)

        self.day_finalized = False
        self.protocol("WM_DELETE_WINDOW", self.on_closing_app)

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.logo_img = ctk.CTkImage(Image.open("logo.png"), size=(80, 80))

        # --- Sidebar ---
        self.sidebar = ctk.CTkFrame(self, width=260, corner_radius=0, fg_color="#1a1c23")
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        ctk.CTkLabel(self.sidebar, text="", image=self.logo_img).pack(pady=(30, 10))
        ctk.CTkLabel(self.sidebar, text="TutorManager Pro+", font=("Montserrat", 20, "bold"),
                     text_color="#0A84FF").pack(pady=(0, 20))

        # --- Sidebar Buttons ---
        self.btn_calendar = ctk.CTkButton(self.sidebar, text="📅 Ημερολόγιο", font=("Arial", 14), height=40,
                                          fg_color="#2980b9", command=self.show_calendar_view)
        self.btn_calendar.pack(pady=5, padx=20, fill="x")

        self.btn_add_student = ctk.CTkButton(self.sidebar, text="➕ Προσθήκη", font=("Arial", 14), height=40,
                                             fg_color="#34495e", command=self.show_add_student_ui)
        self.btn_add_student.pack(pady=5, padx=20, fill="x")

        self.btn_manage_students = ctk.CTkButton(self.sidebar, text="👥 Διαχείριση", font=("Arial", 14), height=40,
                                                 fg_color="#8e44ad", hover_color="#9b59b6",
                                                 command=self.show_manage_students_ui)
        self.btn_manage_students.pack(pady=5, padx=20, fill="x")

        self.btn_export = ctk.CTkButton(self.sidebar, text="📊 Έσοδα (Excel)", font=("Arial", 14), height=40,
                                        fg_color="#27ae60", hover_color="#219150", command=self.export_excel)
        self.btn_export.pack(pady=(20, 5), padx=20, fill="x")

        self.btn_annual = ctk.CTkButton(self.sidebar, text="📈 Ετήσια Αναφορά", font=("Arial", 14), height=40,
                                        fg_color="#16a085", hover_color="#1abc9c", command=self.export_annual_excel)
        self.btn_annual.pack(pady=5, padx=20, fill="x")

        self.btn_backup = ctk.CTkButton(self.sidebar, text="💾 Backup", font=("Arial", 14), height=35,
                                        fg_color="#f39c12", hover_color="#e67e22", text_color="black",
                                        command=self.create_backup)
        self.btn_backup.pack(pady=(30, 5), padx=20, fill="x")

        self.btn_restore = ctk.CTkButton(self.sidebar, text="🔄 Επαναφορά", font=("Arial", 14), height=35,
                                         fg_color="#d35400", hover_color="#e67e22", text_color="white",
                                         command=self.restore_backup)
        self.btn_restore.pack(pady=5, padx=20, fill="x")

        self.btn_summary = ctk.CTkButton(self.sidebar, text="✅ Κλείσιμο Ημέρας", font=("Arial", 15, "bold"), height=45,
                                         fg_color="#c0392b", hover_color="#e74c3c", command=self.show_daily_summary)
        self.btn_summary.pack(pady=20, padx=20, fill="x", side="bottom")

        # Frames
        self.calendar_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.add_student_frame = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.manage_student_frame = ctk.CTkScrollableFrame(self, fg_color="transparent")

        self.setup_add_student_ui()
        self.show_splash()

    # ==================== OΜΑΛΑ ΓΡΑΦΙΚΑ ====================
    def animate_frame_transition(self, target_frame):
        self.hide_all_frames()
        self.after(50, lambda: target_frame.grid(row=0, column=1, sticky="nsew", padx=20, pady=20))

    def hide_all_frames(self):
        self.calendar_frame.grid_forget()
        self.add_student_frame.grid_forget()
        self.manage_student_frame.grid_forget()

    # ==================== SPLASH SCREEN ====================
    def show_splash(self):
        self.splash = ctk.CTkToplevel(self)
        self.splash.title("Φόρτωση...")
        self.splash.geometry("450x300")
        self.splash.overrideredirect(True)

        self.splash.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (450 // 2)
        y = (self.winfo_screenheight() // 2) - (300 // 2)
        self.splash.geometry(f"+{x}+{y}")

        self.splash.configure(fg_color="#1a1c23")
        logo = ctk.CTkImage(Image.open("logo.png"), size=(100, 100))
        ctk.CTkLabel(self.splash, text="", image=logo).pack(pady=(40, 10))
        ctk.CTkLabel(self.splash, text="TUTOR MANAGER PRO+", font=("Montserrat", 22, "bold"),
                     text_color="#0A84FF").pack()

        self.lbl_status = ctk.CTkLabel(self.splash, text="Φόρτωση γραφικών...", font=("Arial", 14), text_color="gray")
        self.lbl_status.pack(pady=20)

        self.after(1000, lambda: self.lbl_status.configure(text="Εκκίνηση..."))
        self.after(2000, self.finish_splash)

    def finish_splash(self):
        self.splash.destroy()
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (1400 // 2)
        y = (self.winfo_screenheight() // 2) - (850 // 2)
        self.geometry(f"+{x}+{y}")
        self.deiconify()
        self.attributes("-alpha", 1.0)
        self.show_calendar_view()

    # ==================== BACKUP & RESTORE ====================
    def create_backup(self):
        try:
            shutil.copy("tutor_manager.db", "tutor_manager_backup.db")
            show_custom_alert(self, "Backup Επιτυχές", "Το αντίγραφο ασφαλείας ανανεώθηκε επιτυχώς!")
        except Exception as e:
            show_custom_alert(self, "Σφάλμα", str(e), is_error=True)

    def restore_backup(self):
        if not os.path.exists("tutor_manager_backup.db"):
            show_custom_alert(self, "Προσοχή", "Δεν βρέθηκε backup.", is_error=True)
            return
        try:
            shutil.copy("tutor_manager_backup.db", "tutor_manager.db")
            show_custom_alert(self, "Επαναφορά Επιτυχής", "Η βάση δεδομένων ανανεώθηκε!")
            self.show_calendar_view()
        except Exception as e:
            show_custom_alert(self, "Σφάλμα", str(e), is_error=True)

    # ==================== ΟΘΟΝΗ ΗΜΕΡΟΛΟΓΙΟΥ (ΑΥΣΤΗΡΑ 50-50) ====================
    def show_calendar_view(self):
        self.animate_frame_transition(self.calendar_frame)
        for w in self.calendar_frame.winfo_children(): w.destroy()

        # Η προσθήκη του uniform="half" ΕΞΑΝΑΓΚΑΖΕΙ το tkinter να τα κάνει ακριβώς 50-50!
        self.calendar_frame.grid_columnconfigure(0, weight=1, uniform="half")
        self.calendar_frame.grid_columnconfigure(1, weight=1, uniform="half")
        self.calendar_frame.grid_rowconfigure(0, weight=1)

        left_panel = ctk.CTkFrame(self.calendar_frame, fg_color="#232731", corner_radius=15)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        cal_header = ctk.CTkFrame(left_panel, fg_color="transparent")
        cal_header.pack(fill="x", pady=15, padx=20)

        current_month = datetime.today().month
        current_year = datetime.today().year

        self.combo_month = ctk.CTkOptionMenu(cal_header, values=[f"{i:02d}" for i in range(1, 13)],
                                             font=("Arial", 16, "bold"), width=90, height=35,
                                             command=lambda _: self.build_calendar_grid())
        self.combo_month.set(f"{current_month:02d}")
        self.combo_month.pack(side="left", padx=10)

        self.combo_year = ctk.CTkOptionMenu(cal_header, values=[str(y) for y in range(2024, 2030)],
                                            font=("Arial", 16, "bold"), width=110, height=35,
                                            command=lambda _: self.build_calendar_grid())
        self.combo_year.set(str(current_year))
        self.combo_year.pack(side="left", padx=10)

        # Το container του grid του ημερολογίου
        self.cal_grid_container = ctk.CTkFrame(left_panel, fg_color="transparent")
        self.cal_grid_container.pack(fill="both", expand=True, padx=20, pady=20)

        for i in range(7):
            self.cal_grid_container.grid_columnconfigure(i, weight=1)
        for i in range(1, 7):
            self.cal_grid_container.grid_rowconfigure(i, weight=1)

        self.right_panel = ctk.CTkFrame(self.calendar_frame, fg_color="#1a1c23", corner_radius=15)
        self.right_panel.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        self.lbl_day_header = ctk.CTkLabel(self.right_panel, text="", font=("Arial", 26, "bold"), text_color="#f1c40f")
        self.lbl_day_header.pack(pady=20)

        # Έδωσα λίγο παραπάνω padding στο δεξί μέρος για να απλώνει όμορφα
        self.scroll_area = ctk.CTkScrollableFrame(self.right_panel, fg_color="transparent")
        self.scroll_area.pack(fill="both", expand=True, padx=25, pady=10)

        self.frame_quick_add = ctk.CTkFrame(self.right_panel, fg_color="#1a1c23", border_width=1,
                                            border_color="#34495e")
        self.frame_quick_add.pack(fill="x", pady=15, padx=25, ipady=10)

        self.build_calendar_grid()
        self.select_day(datetime.today().year, datetime.today().month, datetime.today().day)

    def build_calendar_grid(self):
        for w in self.cal_grid_container.winfo_children(): w.destroy()

        self.cal_buttons = {}
        year = int(self.combo_year.get())
        month = int(self.combo_month.get())

        days_titles = ["Δευ", "Τρι", "Τετ", "Πεμ", "Παρ", "Σαβ", "Κυρ"]
        for i, d in enumerate(days_titles):
            ctk.CTkLabel(self.cal_grid_container, text=d, font=("Arial", 18, "bold"), text_color="#0A84FF").grid(row=0,
                                                                                                                 column=i,
                                                                                                                 pady=10,
                                                                                                                 sticky="nsew")

        cal_matrix = calendar.monthcalendar(year, month)

        with sqlite3.connect("tutor_manager.db") as conn:
            cursor = conn.cursor()
            for r, week in enumerate(cal_matrix):
                for c, day in enumerate(week):
                    if day != 0:
                        date_str = f"{year}-{month:02d}-{day:02d}"
                        cursor.execute("SELECT SUM(hours_done) FROM session_logs WHERE date=?", (date_str,))
                        hours_sum = cursor.fetchone()[0]

                        bg_color = "#27ae60" if hours_sum and hours_sum > 0 else "#2c3e50"
                        is_today = (
                                    day == datetime.today().day and month == datetime.today().month and year == datetime.today().year)
                        border = 3 if is_today else 0
                        border_color = "#f1c40f" if is_today else "gray"

                        text_display = f"{day}\n({hours_sum}h)" if hours_sum else str(day)

                        btn = ctk.CTkButton(self.cal_grid_container, text=text_display,
                                            fg_color=bg_color, font=("Arial", 16, "bold"),
                                            border_width=border, border_color=border_color,
                                            hover_color="#1e8449",
                                            command=lambda d=day, m=month, y=year: self.select_day(y, m, d))
                        btn.grid(row=r + 1, column=c, padx=5, pady=5, sticky="nsew")
                        self.cal_buttons[date_str] = {"btn": btn, "is_today": is_today}

        if hasattr(self, 'selected_date'):
            self.highlight_selected_day()

    def select_day(self, year, month, day):
        self.selected_date = f"{year}-{month:02d}-{day:02d}"
        date_obj = datetime(year, month, day)
        days_gr = ["Δευτέρα", "Τρίτη", "Τετάρτη", "Πέμπτη", "Παρασκευή", "Σάββατο", "Κυριακή"]
        self.selected_day_name = days_gr[date_obj.weekday()]

        self.lbl_day_header.configure(text=f"Ημερήσιος Έλεγχος: {self.selected_day_name} {day:02d}/{month:02d}/{year}")
        self.refresh_day_lists()
        self.setup_quick_add()
        self.highlight_selected_day()

    def highlight_selected_day(self):
        if not hasattr(self, 'cal_buttons'): return
        for d_str, data in self.cal_buttons.items():
            btn = data["btn"]
            is_today = data["is_today"]
            if d_str == getattr(self, 'selected_date', ''):
                btn.configure(border_width=4, border_color="#3498db")
            else:
                btn.configure(border_width=3 if is_today else 0, border_color="#f1c40f" if is_today else "gray")

    def refresh_day_lists(self):
        for w in self.scroll_area.winfo_children(): w.destroy()

        with sqlite3.connect("tutor_manager.db") as conn:
            cursor = conn.cursor()
            ctk.CTkLabel(self.scroll_area, text="📋 Προγραμματισμένα για σήμερα:", font=("Arial", 20, "bold")).pack(
                anchor="w", pady=(0, 15))

            cursor.execute('''SELECT s.id, s.name, s.group_name, s.rate_per_hour, s.hours_per_session 
                              FROM students s JOIN schedule sch ON s.id = sch.student_id WHERE sch.day_of_week = ?''',
                           (self.selected_day_name,))
            scheduled_students = cursor.fetchall()
            scheduled_ids = []

            if not scheduled_students:
                ctk.CTkLabel(self.scroll_area, text="Κανένα προγραμματισμένο μάθημα.", text_color="gray",
                             font=("Arial", 15)).pack(anchor="w", padx=10)
            else:
                for sid, name, gname, rate, hours in scheduled_students:
                    scheduled_ids.append(sid)
                    cursor.execute("SELECT id, hours_done, notes FROM session_logs WHERE student_id=? AND date=?",
                                   (sid, self.selected_date))
                    exists = cursor.fetchone()

                    f = ctk.CTkFrame(self.scroll_area, fg_color="#2b323d", corner_radius=10)
                    f.pack(fill="x", pady=8, ipady=10)
                    title = f"🎓 {name} {'[' + gname + ']' if gname else '[Ατομικό]'}"
                    ctk.CTkLabel(f, text=title, font=("Arial", 16, "bold")).pack(side="left", padx=15)

                    if exists:
                        log_id, hours_logged, notes = exists
                        ctk.CTkButton(f, text="↺ Αναίρεση", fg_color="#e74c3c", width=100, height=35,
                                      font=("Arial", 14, "bold"), hover_color="#c0392b",
                                      command=lambda l=log_id: self.delete_log(l)).pack(side="right", padx=15)
                        ctk.CTkLabel(f, text=f"✔️ {hours_logged}h", text_color="#2ecc71",
                                     font=("Arial", 16, "bold")).pack(side="right", padx=10)
                        if notes:
                            ctk.CTkLabel(f, text=f"📝 {notes}", font=("Arial", 14, "italic"), text_color="gray").pack(
                                side="left", padx=10)
                    else:
                        e_h = ctk.CTkEntry(f, width=65, height=35, font=("Arial", 15))
                        e_h.insert(0, str(hours))
                        e_notes = ctk.CTkEntry(f, width=200, height=35, font=("Arial", 14),
                                               placeholder_text="Σημειώσεις (προαιρετικά)")
                        btn = ctk.CTkButton(f, text="Επιβεβαίωση", fg_color="#27ae60", hover_color="#219150", width=120,
                                            height=35, font=("Arial", 14, "bold"),
                                            command=lambda s=sid, r=rate, e=e_h, n=e_notes: self.save_lesson(s, r, e,
                                                                                                             n))
                        btn.pack(side="right", padx=15)
                        e_h.pack(side="right", padx=10)
                        ctk.CTkLabel(f, text="Ώρες:", font=("Arial", 15)).pack(side="right")
                        e_notes.pack(side="right", padx=20)

            cursor.execute('''SELECT l.id, s.name, s.group_name, l.hours_done, s.id, l.notes 
                              FROM session_logs l JOIN students s ON l.student_id = s.id 
                              WHERE l.date = ?''', (self.selected_date,))
            all_logs_today = cursor.fetchall()

        extra_logs = [log for log in all_logs_today if log[4] not in scheduled_ids]
        if extra_logs:
            ctk.CTkLabel(self.scroll_area, text="⚠️ Έκτακτες Καταχωρήσεις / Αναπληρώσεις:", font=("Arial", 18, "bold"),
                         text_color="#e67e22").pack(anchor="w", pady=(25, 10))
            for lid, name, gname, hrs, _, nots in extra_logs:
                f = ctk.CTkFrame(self.scroll_area, fg_color="#161b22")
                f.pack(fill="x", pady=5, ipady=5)
                n_str = f" | 📝 {nots}" if nots else ""
                ctk.CTkLabel(f, text=f"{name} {'[' + gname + ']' if gname else ''} | {hrs}h {n_str}",
                             font=("Arial", 15)).pack(side="left", padx=15)
                ctk.CTkButton(f, text="🗑️", width=40, height=35, fg_color="#c0392b", hover_color="#a93226",
                              command=lambda log_id=lid: self.delete_log(log_id)).pack(side="right", padx=15)

    def setup_quick_add(self):
        for w in self.frame_quick_add.winfo_children(): w.destroy()

        ctk.CTkButton(self.frame_quick_add, text="➕ Νέα Μαζική Αναπλήρωση / Έκτακτο", font=("Arial", 16, "bold"),
                      height=50, fg_color="#2980b9", hover_color="#1f618d",
                      command=self.open_bulk_extra_session_popup).pack(pady=10, padx=20, fill="x")

    def open_bulk_extra_session_popup(self):
        with sqlite3.connect("tutor_manager.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, name, group_name, rate_per_hour FROM students ORDER BY group_name, name")
            all_students = cursor.fetchall()

        if not all_students:
            show_custom_alert(self, "Πληροφορία", "Δεν υπάρχουν καταχωρημένοι μαθητές στο σύστημα.")
            return

        popup = ctk.CTkToplevel(self)
        popup.title("Μαζική Προσθήκη Εκτάκτων")
        popup.geometry("550x650")
        popup.transient(self)
        popup.grab_set()

        popup.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (550 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (650 // 2)
        popup.geometry(f"+{x}+{y}")

        ctk.CTkLabel(popup, text=f"Αναπληρώσεις για: {self.selected_day_name} {self.selected_date}",
                     font=("Arial", 18, "bold"), text_color="#0A84FF").pack(pady=15)

        scroll = ctk.CTkScrollableFrame(popup, fg_color="#1f252e")
        scroll.pack(fill="both", expand=True, padx=20, pady=10)

        self.bulk_vars = {}
        for sid, name, gname, rate in all_students:
            var = ctk.BooleanVar(value=False)
            self.bulk_vars[sid] = {'var': var, 'rate': rate}
            lbl_text = f"{name} ({gname if gname else 'Ατομικό'})"
            cb = ctk.CTkCheckBox(scroll, text=lbl_text, variable=var, font=("Arial", 16))
            cb.pack(anchor="w", pady=8, padx=15)

        bottom_frame = ctk.CTkFrame(popup, fg_color="#2b323d", corner_radius=10)
        bottom_frame.pack(fill="x", padx=20, pady=15, ipady=15)

        entry_h = ctk.CTkEntry(bottom_frame, width=90, height=40, font=("Arial", 15), placeholder_text="Ώρες")
        entry_h.pack(side="left", padx=15)

        entry_n = ctk.CTkEntry(bottom_frame, width=250, height=40, font=("Arial", 14),
                               placeholder_text="Σημειώσεις (π.χ. Αναπλήρωση)")
        entry_n.pack(side="left", padx=10)

        def save_bulk():
            try:
                hrs_text = entry_h.get()
                if not hrs_text:
                    show_custom_alert(popup, "Προσοχή", "Παρακαλώ εισάγετε τις ώρες.", is_error=True)
                    return
                hrs = float(hrs_text)
                notes = entry_n.get().strip()
                selected_any = False

                with sqlite3.connect("tutor_manager.db") as conn:
                    for s_id, data in self.bulk_vars.items():
                        if data['var'].get():
                            selected_any = True
                            conn.execute(
                                "INSERT INTO session_logs (student_id, date, hours_done, earned_amount, notes) VALUES (?,?,?,?,?)",
                                (s_id, self.selected_date, hrs, hrs * data['rate'], notes))

                if not selected_any:
                    show_custom_alert(popup, "Προσοχή", "Επιλέξτε τουλάχιστον έναν μαθητή τικάροντας το κουτάκι του.",
                                      is_error=True)
                    return

                self.day_finalized = False
                self.refresh_day_lists()
                self.build_calendar_grid()
                popup.destroy()
            except ValueError:
                show_custom_alert(popup, "Σφάλμα", "Παρακαλώ εισάγετε έναν έγκυρο αριθμό ωρών.", is_error=True)

        ctk.CTkButton(bottom_frame, text="✅ Αποθήκευση", fg_color="#27ae60", hover_color="#219150",
                      font=("Arial", 16, "bold"), height=40, command=save_bulk).pack(side="right", padx=15)

    def save_lesson(self, sid, rate, entry_widget, notes_widget):
        try:
            hours = float(entry_widget.get())
            notes = notes_widget.get().strip()
            with sqlite3.connect("tutor_manager.db") as conn:
                conn.execute(
                    "INSERT INTO session_logs (student_id, date, hours_done, earned_amount, notes) VALUES (?,?,?,?,?)",
                    (sid, self.selected_date, hours, hours * rate, notes))
            self.day_finalized = False
            self.refresh_day_lists()
            self.build_calendar_grid()
        except ValueError:
            show_custom_alert(self, "Σφάλμα", "Εισάγετε έγκυρο αριθμό.", is_error=True)

    def delete_log(self, log_id):
        with sqlite3.connect("tutor_manager.db") as conn:
            conn.execute("DELETE FROM session_logs WHERE id=?", (log_id,))
        self.day_finalized = False
        self.refresh_day_lists()
        self.build_calendar_grid()

    # ==================== ΔΙΑΧΕΙΡΙΣΗ & ΕΠΕΞΕΡΓΑΣΙΑ ΜΑΘΗΤΩΝ ====================
    def show_manage_students_ui(self):
        self.animate_frame_transition(self.manage_student_frame)
        for w in self.manage_student_frame.winfo_children(): w.destroy()

        ctk.CTkLabel(self.manage_student_frame, text="Διαχείριση Ενεργών Μαθητών", font=("Arial", 24, "bold"),
                     text_color="#0A84FF").pack(pady=20)

        with sqlite3.connect("tutor_manager.db") as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id, name, group_name, rate_per_hour, hours_per_session FROM students ORDER BY group_name, name")
            students = cursor.fetchall()

        if not students:
            ctk.CTkLabel(self.manage_student_frame, text="Δεν υπάρχουν καταχωρημένοι μαθητές.", text_color="gray",
                         font=("Arial", 16)).pack()
            return

        for s in students:
            sid, name, gname, rate, def_hrs = s
            f = ctk.CTkFrame(self.manage_student_frame, fg_color="#2b323d", corner_radius=10)
            f.pack(fill="x", pady=8, padx=20, ipady=10)

            g_text = f"[{gname}]" if gname else "[Ατομικό]"
            ctk.CTkLabel(f, text=f"👤 {name} {g_text}", font=("Arial", 18, "bold")).pack(side="left", padx=15)
            ctk.CTkLabel(f, text=f"Χρέωση: {rate}€/h", text_color="#f1c40f", font=("Arial", 16)).pack(side="left",
                                                                                                      padx=20)

            ctk.CTkButton(f, text="🗑️", fg_color="#c0392b", hover_color="#a93226", width=45, height=35,
                          command=lambda s_id=sid: self.delete_student(s_id)).pack(side="right", padx=10)
            ctk.CTkButton(f, text="✏️ Επεξεργασία", fg_color="#f39c12", hover_color="#d68910", text_color="black",
                          width=120, height=35, font=("Arial", 14, "bold"),
                          command=lambda s_data=s: self.open_edit_student_popup(s_data)).pack(side="right", padx=10)
            ctk.CTkButton(f, text="📊 Ιστορικό", fg_color="#8e44ad", hover_color="#732d91", width=110, height=35,
                          font=("Arial", 14, "bold"),
                          command=lambda s_id=sid, n=name: self.show_student_history(s_id, n)).pack(side="right",
                                                                                                    padx=10)

    def delete_student(self, student_id):
        try:
            with sqlite3.connect("tutor_manager.db") as conn:
                conn.execute("DELETE FROM schedule WHERE student_id=?", (student_id,))
                conn.execute("DELETE FROM session_logs WHERE student_id=?", (student_id,))
                conn.execute("DELETE FROM students WHERE id=?", (student_id,))
            show_custom_alert(self, "Διαγραφή", "Ο μαθητής διεγράφη επιτυχώς.")
            self.show_manage_students_ui()
        except Exception as e:
            show_custom_alert(self, "Σφάλμα", str(e), is_error=True)

    def open_edit_student_popup(self, student_data):
        sid, name, gname, rate, def_hrs = student_data

        popup = ctk.CTkToplevel(self)
        popup.title(f"Επεξεργασία: {name}")
        popup.geometry("450x550")
        popup.transient(self)
        popup.grab_set()

        popup.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (450 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (550 // 2)
        popup.geometry(f"+{x}+{y}")

        ctk.CTkLabel(popup, text="Επεξεργασία Στοιχείων", font=("Arial", 18, "bold"), text_color="#0A84FF").pack(
            pady=15)

        e_name = ctk.CTkEntry(popup, width=300, height=40, font=("Arial", 14), placeholder_text="Όνομα")
        e_name.insert(0, name)
        e_name.pack(pady=10)

        e_group = ctk.CTkEntry(popup, width=300, height=40, font=("Arial", 14),
                               placeholder_text="Γκρουπ (Κενό για ατομικό)")
        if gname: e_group.insert(0, gname)
        e_group.pack(pady=10)

        e_rate = ctk.CTkEntry(popup, width=300, height=40, font=("Arial", 14), placeholder_text="Χρέωση ανά ώρα")
        e_rate.insert(0, str(rate))
        e_rate.pack(pady=10)

        e_hrs = ctk.CTkEntry(popup, width=300, height=40, font=("Arial", 14), placeholder_text="Συνήθης Διάρκεια")
        e_hrs.insert(0, str(def_hrs))
        e_hrs.pack(pady=10)

        ctk.CTkLabel(popup, text="Ημέρες Μαθήματος:", font=("Arial", 15, "bold")).pack(pady=(15, 5))

        current_days = []
        with sqlite3.connect("tutor_manager.db") as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT day_of_week FROM schedule WHERE student_id=?", (sid,))
            current_days = [row[0] for row in cursor.fetchall()]

        days_frame = ctk.CTkFrame(popup, fg_color="transparent")
        days_frame.pack()
        days_dict = {}
        for d in ["Δευτέρα", "Τρίτη", "Τετάρτη", "Πέμπτη", "Παρασκευή", "Σάββατο", "Κυριακή"]:
            v = ctk.BooleanVar(value=(d in current_days))
            cb = ctk.CTkCheckBox(days_frame, text=d[:3], variable=v, width=60)
            cb.pack(side="left", padx=3)
            days_dict[d] = v

        def save_changes():
            try:
                new_n = e_name.get().strip()
                new_g = e_group.get().strip()
                new_r = float(e_rate.get())
                new_h = float(e_hrs.get())
                sel_days = [d for d, v in days_dict.items() if v.get()]

                if not new_n or not sel_days:
                    show_custom_alert(popup, "Προσοχή", "Το όνομα και τουλάχιστον μία μέρα είναι υποχρεωτικά.",
                                      is_error=True)
                    return

                with sqlite3.connect("tutor_manager.db") as conn:
                    conn.execute(
                        "UPDATE students SET name=?, group_name=?, rate_per_hour=?, hours_per_session=? WHERE id=?",
                        (new_n, new_g, new_r, new_h, sid))
                    conn.execute("DELETE FROM schedule WHERE student_id=?", (sid,))
                    for d in sel_days:
                        conn.execute("INSERT INTO schedule (student_id, day_of_week) VALUES (?,?)", (sid, d))

                popup.destroy()
                self.show_manage_students_ui()
                self.build_calendar_grid()
            except ValueError:
                show_custom_alert(popup, "Σφάλμα", "Ελέγξτε τα νούμερα σε Χρέωση/Ώρες.", is_error=True)

        ctk.CTkButton(popup, text="Αποθήκευση", fg_color="#27ae60", hover_color="#219150", font=("Arial", 15, "bold"),
                      height=40, command=save_changes).pack(pady=25)

    def show_student_history(self, student_id, student_name):
        history_win = ctk.CTkToplevel(self)
        history_win.title(f"Ιστορικό: {student_name}")
        history_win.geometry("650x700")
        history_win.transient(self)
        history_win.grab_set()

        history_win.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (650 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (700 // 2)
        history_win.geometry(f"+{x}+{y}")

        ctk.CTkLabel(history_win, text=f"Αναλυτική Καρτέλα: {student_name}", font=("Arial", 22, "bold"),
                     text_color="#0A84FF").pack(pady=15)

        filter_frame = ctk.CTkFrame(history_win, fg_color="transparent")
        filter_frame.pack(pady=5)
        combo_month = ctk.CTkOptionMenu(filter_frame, values=[f"{i:02d}" for i in range(1, 13)], width=90, height=35,
                                        font=("Arial", 14, "bold"))
        combo_month.set(f"{datetime.today().month:02d}")
        combo_month.pack(side="left", padx=5)
        combo_year = ctk.CTkOptionMenu(filter_frame, values=[str(y) for y in range(2024, 2030)], width=110, height=35,
                                       font=("Arial", 14, "bold"))
        combo_year.set(str(datetime.today().year))
        combo_year.pack(side="left", padx=5)

        scroll_area = ctk.CTkScrollableFrame(history_win, fg_color="#1a1c23")
        scroll_area.pack(fill="both", expand=True, padx=20, pady=10)

        totals_frame = ctk.CTkFrame(history_win, fg_color="#2b323d", corner_radius=10)
        totals_frame.pack(fill="x", padx=20, pady=15, ipady=15)
        lbl_total_hours = ctk.CTkLabel(totals_frame, text="Σύνολο Ωρών: 0", font=("Arial", 18, "bold"))
        lbl_total_hours.pack(side="left", padx=25)
        lbl_total_money = ctk.CTkLabel(totals_frame, text="Σύνολο Ποσού: 0.00€", font=("Arial", 18, "bold"),
                                       text_color="#27ae60")
        lbl_total_money.pack(side="right", padx=25)

        def load_data(*args):
            for w in scroll_area.winfo_children(): w.destroy()
            search_date = f"{combo_year.get()}-{combo_month.get()}-%"
            with sqlite3.connect("tutor_manager.db") as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT date, hours_done, earned_amount, notes FROM session_logs WHERE student_id=? AND date LIKE ? ORDER BY date ASC",
                    (student_id, search_date))
                records = cursor.fetchall()

            if not records:
                ctk.CTkLabel(scroll_area, text="Δεν βρέθηκαν μαθήματα.", text_color="gray", font=("Arial", 16)).pack(
                    pady=30)
                lbl_total_hours.configure(text="Σύνολο Ωρών: 0")
                lbl_total_money.configure(text="Σύνολο Ποσού: 0.00€")
                return

            t_h = t_m = 0.0
            for r_date, hrs, mon, nots in records:
                t_h += hrs
                t_m += mon
                d_str = datetime.strptime(r_date, "%Y-%m-%d").strftime("%d/%m/%Y")
                row = ctk.CTkFrame(scroll_area, fg_color="#232731")
                row.pack(fill="x", pady=4, ipady=8)
                ctk.CTkLabel(row, text=f"📅 {d_str}", font=("Arial", 16, "bold")).pack(side="left", padx=15)
                n_str = f" | 📝 {nots}" if nots else ""
                ctk.CTkLabel(row, text=f"{hrs}h{n_str}", text_color="#bdc3c7", font=("Arial", 15)).pack(side="left",
                                                                                                        padx=10)
                ctk.CTkLabel(row, text=f"{mon:.2f}€", font=("Arial", 16, "bold"), text_color="#f1c40f").pack(
                    side="right", padx=15)

            lbl_total_hours.configure(text=f"Σύνολο Ωρών: {t_h}")
            lbl_total_money.configure(text=f"Σύνολο Ποσού: {t_m:.2f}€")

        combo_month.configure(command=load_data)
        combo_year.configure(command=load_data)
        load_data()

    # ==================== ΠΡΟΣΘΗΚΗ ΜΑΘΗΤΩΝ ====================
    def setup_add_student_ui(self):
        card = ctk.CTkFrame(self.add_student_frame, corner_radius=15, fg_color="#232731")
        card.pack(pady=20, padx=40, fill="both", expand=True)
        ctk.CTkLabel(card, text="Προσθήκη Νέου Μαθητή / Γκρουπ", font=("Arial", 24, "bold")).pack(pady=25)

        self.group_var = ctk.BooleanVar(value=False)
        self.cb_is_group = ctk.CTkCheckBox(card, text="Σε Γκρουπ;", font=("Arial", 16), variable=self.group_var,
                                           command=self.toggle_group_fields)
        self.cb_is_group.pack(pady=10)

        self.entry_group = ctk.CTkEntry(card, placeholder_text="Όνομα Γκρουπ (π.χ. Β' Λυκείου)", width=400, height=45,
                                        font=("Arial", 15))
        self.students_container = ctk.CTkFrame(card, fg_color="transparent")
        self.students_container.pack(pady=10)
        self.student_entries = []
        self.add_student_field()

        self.btn_add_more = ctk.CTkButton(card, text="+ Νέο Μέλος Γκρουπ", fg_color="#555", hover_color="#444",
                                          height=40, font=("Arial", 14), command=self.add_student_field)
        self.entry_rate = ctk.CTkEntry(card, placeholder_text="Χρέωση ανά ώρα ανά άτομο (€)", width=400, height=45,
                                       font=("Arial", 15))
        self.entry_rate.pack(pady=10)
        self.entry_def_hours = ctk.CTkEntry(card, placeholder_text="Συνήθης διάρκεια (π.χ. 1.5)", width=400, height=45,
                                            font=("Arial", 15))
        self.entry_def_hours.pack(pady=10)

        ctk.CTkLabel(card, text="Επιλογή Ημερών:", font=("Arial", 18, "bold")).pack(pady=(25, 10))
        self.days_frame = ctk.CTkFrame(card, fg_color="transparent")
        self.days_frame.pack(pady=5)
        self.days_dict_add = {}
        for d in ["Δευτέρα", "Τρίτη", "Τετάρτη", "Πέμπτη", "Παρασκευή", "Σάββατο", "Κυριακή"]:
            v = ctk.BooleanVar()
            cb = ctk.CTkCheckBox(self.days_frame, text=d[:3], variable=v, width=65, font=("Arial", 14))
            cb.pack(side="left", padx=6)
            self.days_dict_add[d] = v

        ctk.CTkButton(card, text="Αποθήκευση", fg_color="#27ae60", hover_color="#219150", height=50,
                      font=("Arial", 16, "bold"), command=self.save_student_to_db).pack(pady=35)

    def toggle_group_fields(self):
        if self.group_var.get():
            self.entry_group.pack(after=self.cb_is_group, pady=10)
            self.btn_add_more.pack(after=self.students_container, pady=5)
        else:
            self.entry_group.pack_forget()
            self.btn_add_more.pack_forget()
            while len(self.student_entries) > 1: self.student_entries.pop().destroy()

    def add_student_field(self):
        en = ctk.CTkEntry(self.students_container, placeholder_text=f"Όνομα {len(self.student_entries) + 1}", width=400,
                          height=45, font=("Arial", 15))
        en.pack(pady=5)
        self.student_entries.append(en)

    def save_student_to_db(self):
        g_name = self.entry_group.get().strip() if self.group_var.get() else ""
        rate, hrs = self.entry_rate.get(), self.entry_def_hours.get()
        s_days = [d for d, v in self.days_dict_add.items() if v.get()]
        names = [e.get().strip() for e in self.student_entries if e.get().strip()]

        if not names or not rate or not s_days:
            show_custom_alert(self, "Προσοχή", "Συμπληρώστε Όνομα, Χρέωση και τουλάχιστον 1 ημέρα.", is_error=True)
            return

        try:
            with sqlite3.connect("tutor_manager.db") as conn:
                for n in names:
                    cursor = conn.cursor()
                    cursor.execute(
                        "INSERT INTO students (name, group_name, rate_per_hour, hours_per_session) VALUES (?,?,?,?)",
                        (n, g_name, float(rate), float(hrs)))
                    sid = cursor.lastrowid
                    for d in s_days: conn.execute("INSERT INTO schedule (student_id, day_of_week) VALUES (?,?)",
                                                  (sid, d))

            show_custom_alert(self, "Επιτυχία", "Η αποθήκευση ολοκληρώθηκε!")
            self.entry_group.delete(0, 'end')
            self.entry_rate.delete(0, 'end')
            self.entry_def_hours.delete(0, 'end')
            for v in self.days_dict_add.values(): v.set(False)
            for e in self.student_entries: e.delete(0, 'end')
        except ValueError:
            show_custom_alert(self, "Σφάλμα", "Μη έγκυρα νούμερα σε χρέωση/ώρες.", is_error=True)

    def show_add_student_ui(self):
        self.animate_frame_transition(self.add_student_frame)

    # ==================== POP-UP ΚΛΕΙΣΙΜΑΤΟΣ ΗΜΕΡΑΣ ====================
    def show_daily_summary(self, is_closing=False):
        today_date = datetime.today().strftime('%Y-%m-%d')
        with sqlite3.connect("tutor_manager.db") as conn:
            cursor = conn.cursor()
            cursor.execute(
                '''SELECT s.name, s.group_name, l.hours_done FROM session_logs l JOIN students s ON l.student_id = s.id WHERE l.date = ?''',
                (today_date,))
            completed = cursor.fetchall()

        if not completed:
            if is_closing:
                self.destroy()
            else:
                show_custom_alert(self, "Πληροφορία", "Δεν υπάρχουν μαθήματα σήμερα.")
            return

        popup = ctk.CTkToplevel(self)
        popup.title("Έλεγχος Ημέρας")
        popup.geometry("550x550")
        popup.transient(self)
        popup.grab_set()

        popup.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (550 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (550 // 2)
        popup.geometry(f"+{x}+{y}")

        if is_closing:
            ctk.CTkLabel(popup, text="⚠️ ΠΡΟΣΟΧΗ!", font=("Arial", 20, "bold"), text_color="#e74c3c").pack(pady=(20, 0))
            ctk.CTkLabel(popup, text="Δεν έχεις επιβεβαιώσει το κλείσιμο ημέρας.", font=("Arial", 14)).pack(pady=5)
        else:
            ctk.CTkLabel(popup, text="📋 Σύνοψη Σήμερα", font=("Arial", 22, "bold"), text_color="#0A84FF").pack(pady=20)

        scroll_popup = ctk.CTkScrollableFrame(popup, fg_color="#1f252e")
        scroll_popup.pack(fill="both", expand=True, padx=20, pady=10)

        for n, gn, hrs in completed:
            row = ctk.CTkFrame(scroll_popup, fg_color="transparent")
            row.pack(fill="x", pady=5)
            ctk.CTkLabel(row, text=f"{n} {'[' + gn + ']' if gn else ''}", font=("Arial", 15)).pack(side="left")
            ctk.CTkLabel(row, text=f"{hrs}h", font=("Arial", 15, "bold"), text_color="#f1c40f").pack(side="right")

        btn_frame = ctk.CTkFrame(popup, fg_color="transparent")
        btn_frame.pack(pady=20)
        ctk.CTkButton(btn_frame, text="🔙 Πίσω", fg_color="#7f8c8d", width=120, command=popup.destroy).pack(side="left",
                                                                                                           padx=10)
        ctk.CTkButton(btn_frame, text="✅ Αποθήκευση & Έξοδος", fg_color="#27ae60", width=160,
                      command=lambda: self.finalize_day(popup, is_closing)).pack(side="left", padx=10)

    def finalize_day(self, popup, is_closing):
        self.day_finalized = True
        popup.destroy()
        if is_closing:
            self.destroy()
        else:
            show_custom_alert(self, "Μπράβο!", "Η μέρα επιβεβαιώθηκε επιτυχώς.")

    def on_closing_app(self):
        with sqlite3.connect("tutor_manager.db") as conn:
            logs = conn.execute("SELECT COUNT(*) FROM session_logs WHERE date=?",
                                (datetime.today().strftime('%Y-%m-%d'),)).fetchone()[0]
        if logs > 0 and not self.day_finalized:
            self.show_daily_summary(is_closing=True)
        else:
            self.destroy()

    # ==================== EXCEL EXPORT (ΜΗΝΙΑΙΟ ΕΣΟΔΩΝ) ====================
    def export_excel(self):
        with sqlite3.connect("tutor_manager.db") as conn:
            query = '''SELECT s.name AS 'Μαθητής', IFNULL(s.group_name, 'Ατομικό') AS 'Γκρουπ', strftime('%Y-%m', l.date) AS 'Μήνας', SUM(l.hours_done) AS 'Σύνολο Ωρών', SUM(l.earned_amount) AS 'Οφειλόμενο Ποσό' FROM session_logs l JOIN students s ON l.student_id = s.id GROUP BY s.id, strftime('%Y-%m', l.date) ORDER BY 'Μήνας' DESC, 'Γκρουπ', 'Μαθητής' '''
            df = pd.read_sql_query(query, conn)

        if df.empty:
            show_custom_alert(self, "Άδειο", "Δεν υπάρχουν δεδομένα μαθημάτων.", is_error=True)
            return

        filename = "Αναφορά_Εσόδων.xlsx"
        self._write_to_excel(df, filename)

    # ==================== EXCEL EXPORT (ΕΤΗΣΙΟ ΣΥΝΟΛΟ ΩΡΩΝ) ====================
    def export_annual_excel(self):
        current_year = str(datetime.today().year)
        with sqlite3.connect("tutor_manager.db") as conn:
            query = f'''SELECT s.name AS 'Μαθητής', IFNULL(s.group_name, 'Ατομικό') AS 'Γκρουπ', strftime('%m', l.date) AS 'Μήνας', l.hours_done AS 'Ώρες' FROM session_logs l JOIN students s ON l.student_id = s.id WHERE strftime('%Y', l.date) = '{current_year}' '''
            df = pd.read_sql_query(query, conn)

        if df.empty:
            show_custom_alert(self, "Άδειο", f"Δεν υπάρχουν δεδομένα για το {current_year}.", is_error=True)
            return

        df_pivot = df.pivot_table(index=['Μαθητής', 'Γκρουπ'], columns='Μήνας', values='Ώρες', aggfunc='sum',
                                  fill_value=0).reset_index()
        months_cols = [c for c in df_pivot.columns if c not in ['Μαθητής', 'Γκρουπ']]
        df_pivot['Σύνολο Έτους'] = df_pivot[months_cols].sum(axis=1)

        filename = f"Ετήσια_Αναφορά_Ωρών_{current_year}.xlsx"
        self._write_to_excel(df_pivot, filename, is_annual=True)

    def _write_to_excel(self, df, filename, is_annual=False):
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Αναλυτικά')
                ws = writer.sheets['Αναλυτικά']

                header_fill = PatternFill(start_color="0A84FF", end_color="0A84FF", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                bottom=Side(style='thin'))

                for cell in ws["1:1"]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = border
                        if not is_annual and cell.column_letter == 'E':
                            cell.number_format = '#,##0.00 €'
                        elif is_annual and cell.column not in [1, 2]:
                            cell.alignment = Alignment(horizontal="center")

                for col in ws.columns:
                    col_letter = col[0].column_letter
                    max_len = max([len(str(c.value)) for c in col if c.value] + [0])
                    ws.column_dimensions[col_letter].width = max_len + 5

            show_custom_alert(self, "Ολοκληρώθηκε", f"Δημιουργήθηκε το αρχείο:\n{filename}")
            filepath = os.path.abspath(filename)
            if platform.system() == 'Windows':
                os.startfile(filepath)
            elif platform.system() == 'Darwin':
                os.system(f"open '{filepath}'")
            else:
                os.system(f"xdg-open '{filepath}'")
        except Exception as e:
            show_custom_alert(self, "Σφάλμα Excel", str(e), is_error=True)


# --- ΕΚΚΙΝΗΣΗ ---
if __name__ == "__main__":
    ensure_logo_exists()
    init_db()
    app = TutorApp()
    app.mainloop()